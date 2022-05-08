#!/usr/bin/env python

# Simple script to read the output from OpenVASReporting and
# add in some extra fields


# https://www.softwaretestinghelp.com/python-openpyxl-tutorial/
# https://pythoninoffice.com/adjust-excel-fonts-using-python-openpyxl/

import openpyxl
import re
import sys
import argparse
import os

from openpyxl.styles import Font
from copy import copy
import xml.etree.ElementTree as et


def main():

    parser = argparse.ArgumentParser()  
    parser.add_argument('-i' , '--input-xlsx',  type=argparse.FileType('r'), help="Input name of xlsx to read" )
    parser.add_argument('-o' ,  '--output-xlsx', help="Input name of xlsx to output" )
    parser.add_argument('xmlfile', type=argparse.FileType('r'), nargs='+', help='The xml files to read')

    args = parser.parse_args()

    # Check is import file exists
    if not os.path.exists( args.input_xlsx.name ):
        print( "The input file {0} does not exist".format( args.input_xlsx.name  ))
        sys.exit(1)

     # Check is ouput file exists
    if os.path.exists( args.output_xlsx ):
        print( "The output file {0} already exists".format( args.output_xlsx  ))
        sys.exit(1)    


    # Lets open the main xlsx import file
    try:
        wb = openpyxl.load_workbook( args.input_xlsx.name )
    except:
        print("Something went wrong trying to open xlsx input file")
        raise Exception 


    # Lets open and parse the  XML Files
    section_info = {}
    for xml_file in args.xmlfile:
        root = et.parse( xml_file.name ).getroot()
        tag = root.tag
       

        # This feels wrong. Lets open the supplied XML files and scan them for the CVE Dates.
        # I think an alternative to this would be some kind of external lookup.

        section_name = None
        section_date = None
        for report in root.findall('report'):
            for results in report.findall('results'):
                for reports in results:
                    for i in reports:
                        if i.tag == 'name':
                            print( i.text )
                            section_name = i.text 
                        if i.tag == 'nvt':
                            for nvt in i:
                                if nvt.tag == 'severities':
                                    for severities in nvt:
                                        if severities.tag == 'severity':
                                            for severity in severities:
                                                if severity.tag == 'date':
                                                    section_date = severity.text
                                                    section_info[ section_name ] = section_date

                 

    print(section_info )

    toc = wb["TOC"]


    # Set title
    print("Setting value")
    toc.cell(column=6, row=3).value = 'CVE Name'
    toc.cell(column=7, row=3).value = 'CVE Date'

    # Take a copy of the formatting alraedy used and apply to the new header we have created.
    print("Setting font by coping column=2, row=3")
    toc.cell(column=6, row=3).font  = copy( toc.cell(column=2, row=3).font  )
    toc.cell(column=7, row=3).font  = copy( toc.cell(column=2, row=3).font  )

    print("Setting full by coping column=2, row=3")
    toc.cell(column=6, row=3).fill  = copy( toc.cell(column=2, row=3).fill  )
    toc.cell(column=7, row=3).fill  = copy( toc.cell(column=2, row=3).fill  )

    # Work through rows in TOC. Lets start at row 4
    row=4
    while True:
        # Take the  ID and Name from a the current row ( row )
        id = toc.cell(column=2, row=row)
        name = toc.cell(column=3, row=row)

        # Id thing is set, then we are at the end.
        if id.value == None:
            break

        if name.value == None:
            break

        # Check if a work book exits starting with the id
        match_sheet = None
        for i in wb.sheetnames:
            if i.startswith( id.value + '_'):
                match_sheet = i
                continue 


        # If we do find a matching sheet ( match_sheet is not set) then break
        if not match_sheet:
            print("Hulk sad. No matching work sheet for ", title)
            break


        # Load worksheet 
        print("Loading worksheet ", match_sheet)
        worksheet = wb[ match_sheet ]
        cves = worksheet.cell( column=3, row=7).value
        
        # Write to TOC
        print("Writing CVE to TOC col=6 row=",row)
        toc.cell(column=6, row=row).value = cves

        # See if we can match the CVE name to the date info retrieved from the XML Search. 
        # If so write the data
        if section_info[ name.value  ]:
            print("Writing DATE to TOC col=7 row=",row)
            toc.cell(column=7, row=row).value = section_info[ name.value  ]

        # Inc for next row
        row = row + 1


    print("Saving ", args.output_xlsx)
    wb.save(args.output_xlsx)



if __name__ == '__main__':
    main()   